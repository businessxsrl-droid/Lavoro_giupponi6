"""
Flask Backend — Lavoro Giupponi6
API completa per il frontend Calor Systems Riconciliazione.
Serve il frontend da Lavoro_Giupponi5/frontend.
"""
import os
import sys
import hashlib
import datetime
import io
import json
import urllib.request
import urllib.error
from collections import Counter

import pandas as pd
from flask import Flask, request, jsonify, render_template, send_from_directory, send_file
from flask_jwt_extended import (
    JWTManager, create_access_token, create_refresh_token,
    jwt_required, get_jwt_identity,
)
from flask_cors import CORS

# ── Path setup ────────────────────────────────────────────────────────────────
ROOT_DIR      = os.path.dirname(os.path.abspath(__file__))
FRONTEND_DIR  = os.path.join(ROOT_DIR, "frontend")
TEMPLATES_DIR = os.path.join(FRONTEND_DIR, "templates")
STATIC_DIR    = os.path.join(FRONTEND_DIR, "static")
EXCEL_DIR     = os.path.join(ROOT_DIR, "excel")

# ── Imports locali ────────────────────────────────────────────────────────────
from database   import get_connection, init_db, get_config
from classifier import identify_file_type
from ingestion  import ingest_folder
from reconciler import reconcile, _calcola_stato

# ── Inizializzazione DB ───────────────────────────────────────────────────────
init_db()
os.makedirs(EXCEL_DIR, exist_ok=True)

# ── Flask app ─────────────────────────────────────────────────────────────────
app = Flask(__name__, template_folder=TEMPLATES_DIR, static_folder=STATIC_DIR)
app.config["JWT_SECRET_KEY"]             = os.environ.get("JWT_SECRET", "calor-systems-secret-2025")
app.config["JWT_ACCESS_TOKEN_EXPIRES"]   = datetime.timedelta(hours=8)
app.config["JWT_REFRESH_TOKEN_EXPIRES"]  = datetime.timedelta(days=7)

CORS(app)
jwt = JWTManager(app)


# ═══════════════════════════════════════════════════════════════════════════════
#  TEMPLATE / STATIC
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/login")
def login_page():
    return render_template("login.html")

@app.route("/static/<path:path>")
def send_static(path):
    return send_from_directory(STATIC_DIR, path)


# ═══════════════════════════════════════════════════════════════════════════════
#  AUTH
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/auth/login", methods=["POST"])
def login():
    data     = request.get_json(force=True)
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""

    conn = get_connection()
    row  = conn.execute(
        "SELECT password_hash FROM users WHERE username = ?", (username,)
    ).fetchone()
    conn.close()

    if row and row["password_hash"] == hashlib.sha256(password.encode()).hexdigest():
        access  = create_access_token(identity=username)
        refresh = create_refresh_token(identity=username)
        return jsonify(access_token=access, refresh_token=refresh), 200

    return jsonify(msg="Credenziali non valide"), 401


@app.route("/api/auth/refresh", methods=["POST"])
@jwt_required(refresh=True)
def refresh_token():
    identity = get_jwt_identity()
    access   = create_access_token(identity=identity)
    return jsonify(access_token=access), 200


# ═══════════════════════════════════════════════════════════════════════════════
#  CLASSIFY — identifica un singolo file senza salvarlo
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/classify", methods=["POST"])
@jwt_required()
def classify():
    if "file" not in request.files:
        return jsonify(error="Nessun file inviato"), 400

    f         = request.files["file"]
    tmp_path  = os.path.join(ROOT_DIR, "tmp_" + f.filename)
    f.save(tmp_path)

    try:
        res = identify_file_type(tmp_path)
        return jsonify(res)
    except Exception as e:
        return jsonify(categoria="ERRORE", confidenza=0, ragione=str(e)), 500
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except OSError:
            pass


# ═══════════════════════════════════════════════════════════════════════════════
#  UPLOAD + PIPELINE COMPLETA
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/upload", methods=["POST"])
@jwt_required()
def upload():
    files = request.files.getlist("files[]")
    if not files:
        return jsonify(error="Nessun file selezionato"), 400

    os.makedirs(EXCEL_DIR, exist_ok=True)

    # Rimuovi i file Excel precedenti così si elaborano SOLO i file appena caricati
    for old_fn in os.listdir(EXCEL_DIR):
        if old_fn.lower().endswith((".xlsx", ".xls")) and not old_fn.startswith("~$"):
            try:
                os.remove(os.path.join(EXCEL_DIR, old_fn))
            except OSError:
                pass

    saved = 0
    for f in files:
        if f.filename and not f.filename.startswith("~$"):
            dest = os.path.join(EXCEL_DIR, f.filename)
            f.save(dest)
            saved += 1

    if saved == 0:
        return jsonify(error="Nessun file salvato"), 400

    logs = [f"Salvati {saved} file in {EXCEL_DIR}"]

    # ── Ingestion ─────────────────────────────────────────────────────────────
    try:
        summary = ingest_folder(EXCEL_DIR)
        logs.append(f"Fortech: {summary.get('FORTECH', 0)} righe")
        logs.append(f"Contanti: {summary.get('CONTANTI', 0)} righe")
        logs.append(f"Carte Bancarie: {summary.get('CARTE_BANCARIE', 0)} righe")
        logs.append(f"Satispay: {summary.get('SATISPAY', 0)} righe")
        logs.append(f"Buoni: {summary.get('BUONI', 0)} righe")
        logs.append(f"Carte Petrolifere: {summary.get('carte_petrolifere', 0)} righe")
    except Exception as e:
        logs.append(f"[ERR] Ingestion: {e}")
        return jsonify(error=str(e), logs=logs), 500

    # ── Riconciliazione ───────────────────────────────────────────────────────
    try:
        n_ric = reconcile()
        logs.append(f"Riconciliazione: {n_ric} record elaborati")
    except Exception as e:
        logs.append(f"[WARN] Riconciliazione: {e}")

    # Conta giornate elaborate
    conn = get_connection()
    days = (conn.execute(
        "SELECT COUNT(DISTINCT data) FROM transazioni_fortech").fetchone() or [0])[0]
    conn.close()

    return jsonify({
        "files_imported": saved,
        "days_analyzed":  days,
        "status":         "success",
        "logs":           logs,
    })


# ═══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD — STATS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/stats", methods=["GET"])
@jwt_required()
def get_stats():
    conn = get_connection()

    def _count(q):
        row = conn.execute(q).fetchone()
        return row[0] if row else 0

    total_impianti  = _count("SELECT COUNT(*) as n FROM impianti")
    total_giornate  = _count("SELECT COUNT(DISTINCT data) as n FROM transazioni_fortech")
    quadrate        = _count("SELECT COUNT(*) as n FROM riconciliazione_risultati "
                             "WHERE stato IN ('QUADRATO','QUADRATO_ARROT')")
    anomalie        = _count("SELECT COUNT(*) as n FROM riconciliazione_risultati "
                             "WHERE stato IN ('ANOMALIA_LIEVE','ANOMALIA_GRAVE','NON_TROVATO')")
    anomalie_gravi  = _count("SELECT COUNT(*) as n FROM riconciliazione_risultati "
                             "WHERE stato = 'ANOMALIA_GRAVE'")
    fortech_records = _count("SELECT COUNT(*) as n FROM transazioni_fortech")

    conn.close()
    return jsonify({
        "total_impianti":  total_impianti,
        "total_giornate":  total_giornate,
        "quadrate":        quadrate,
        "anomalie_aperte": anomalie,
        "anomalie_gravi":  anomalie_gravi,
        "fortech_records": fortech_records,
    })


# ═══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD — STATO VERIFICHE PER IMPIANTO
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/stato-verifiche", methods=["GET"])
@jwt_required()
def get_stato_verifiche():
    conn = get_connection()

    impianti = conn.execute(
        "SELECT id, codice_pv, nome, tipo_gestione FROM impianti ORDER BY nome").fetchall()

    results = []
    for imp in impianti:
        rows = conn.execute('''
            SELECT categoria, stato
            FROM riconciliazione_risultati
            WHERE codice_pv = ?
            ORDER BY data DESC
        ''', (imp["codice_pv"],)).fetchall()

        # Per ogni categoria: prendi lo stato peggiore più recente
        cat_stato: dict[str, str] = {}
        _priority = {
            "ANOMALIA_GRAVE": 5, "ANOMALIA_LIEVE": 4, "NON_TROVATO": 3,
            "IN_ATTESA": 2, "QUADRATO_ARROT": 1, "QUADRATO": 0,
        }
        for r in rows:
            cat = r["categoria"]
            s   = r["stato"]
            if cat not in cat_stato:
                cat_stato[cat] = s
            elif _priority.get(s, 0) > _priority.get(cat_stato[cat], 0):
                cat_stato[cat] = s

        results.append({
            "id":            imp["id"],
            "nome":          imp["nome"] or f"PV {imp['codice_pv']}",
            "tipo_gestione": imp["tipo_gestione"] or "PRESIDIATO",
            "categorie":     {cat: {"stato": stato} for cat, stato in cat_stato.items()},
        })

    conn.close()
    return jsonify(results)


# ═══════════════════════════════════════════════════════════════════════════════
#  RICONCILIAZIONI — TABELLA
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/riconciliazioni", methods=["GET"])
@jwt_required()
def get_riconciliazioni():
    da = request.args.get("da")
    a  = request.args.get("a")
    pv = request.args.get("pv")

    query  = '''
        SELECT r.id, r.codice_pv, r.data, r.categoria,
               r.valore_teorico, r.valore_reale, r.differenza, r.stato, r.note, r.tipo_match,
               i.nome AS nome_pv
        FROM riconciliazione_risultati r
        LEFT JOIN impianti i ON r.codice_pv = i.codice_pv
        WHERE 1=1
    '''
    params = []
    if da:
        query  += " AND r.data >= ?"
        params.append(da)
    if a:
        query  += " AND r.data <= ?"
        params.append(a)
    if pv:
        query  += " AND r.codice_pv = ?"
        params.append(int(pv))
    query += " ORDER BY r.data DESC, r.codice_pv, r.categoria"

    conn = get_connection()
    rows = conn.execute(query, params).fetchall()
    conn.close()

    # Etichette leggibili per le categorie
    _CAT_LABEL = {
        "carte_bancarie":          "Carte bancarie (POS)",
        "satispay":                "Satispay",
        "buoni":                   "Buoni / Voucher",
        "carte_petrolifere":       "Carte petrolifere",
        "buoni_petrolifere":       "Buoni + Petrolifere",
        "buoni_petrolifere_combined": "Buoni + Petrolifere (combinato)",
        "prove_erogazione":        "Prove di erogazione",
        "clienti_fine_mese":       "Clienti fine mese",
        "diversi":                 "Diversi",
    }

    return jsonify([{
        "id":               r["id"],
        "data":             r["data"],
        "impianto":         f"{r['codice_pv']} – {r['nome_pv'] or 'N/D'}",
        "categoria":        r["categoria"],
        "categoria_label":  _CAT_LABEL.get(r["categoria"], r["categoria"]),
        "valore_fortech":   r["valore_teorico"],
        "valore_reale":     r["valore_reale"],
        "differenza":       r["differenza"],
        "stato":            r["stato"],
        "tipo_match":       r["tipo_match"] or "nessuno",
        "note":             r["note"] or "",
    } for r in rows])


# ═══════════════════════════════════════════════════════════════════════════════
#  RICONCILIAZIONI — MODIFICA INLINE
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/riconciliazioni/edit", methods=["POST"])
@jwt_required()
def edit_riconciliazione():
    data = request.get_json(force=True)
    rid          = data.get("id")
    valore_reale = data.get("valore_reale")
    note         = data.get("note", "")

    if rid is None:
        return jsonify(error="id mancante"), 400

    conn = get_connection()
    row  = conn.execute(
        "SELECT valore_teorico FROM riconciliazione_risultati WHERE id = ?", (rid,)
    ).fetchone()

    if not row:
        conn.close()
        return jsonify(error="Record non trovato"), 404

    try:
        valore_reale = float(valore_reale)
    except (TypeError, ValueError):
        conn.close()
        return jsonify(error="valore_reale non valido"), 400

    teorico    = float(row["valore_teorico"])
    differenza = round(teorico - valore_reale, 2)

    cfg       = get_config(conn)
    # Determina categoria per scegliere la tolleranza giusta
    cat_row   = conn.execute(
        "SELECT categoria FROM riconciliazione_risultati WHERE id = ?", (rid,)
    ).fetchone()
    cat       = cat_row["categoria"] if cat_row else "carte_bancarie"
    tol_map   = {
        "carte_bancarie": float(cfg.get("tolleranza_carte_fisiologica",       1.0)),
        "satispay":       float(cfg.get("tolleranza_satispay",                0.01)),
        "buoni":          float(cfg.get("tolleranza_buoni",                   0.01)),
        "carte_petrolifere": float(cfg.get("tolleranza_petrolifere",             0.01)),
    }
    tol = tol_map.get(cat, 0.01)

    nuovo_stato = _calcola_stato(teorico, valore_reale, tol) or "IN_ATTESA"

    conn.execute('''
        UPDATE riconciliazione_risultati
        SET valore_reale = ?, differenza = ?, stato = ?, note = ?
        WHERE id = ?
    ''', (valore_reale, differenza, nuovo_stato, note, rid))
    conn.commit()
    conn.close()

    return jsonify(differenza=differenza, nuovo_stato=nuovo_stato)


# ═══════════════════════════════════════════════════════════════════════════════
#  RICONCILIAZIONI — EXPORT EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/riconciliazioni/export/excel", methods=["GET"])
@jwt_required()
def export_excel():
    da = request.args.get("da")
    a  = request.args.get("a")
    pv = request.args.get("pv")

    # Nessuna colonna extra ripetuta: prove_erogazione, clienti_fine_mese, diversi
    # sono ora righe di categoria in riconciliazione_risultati come tutte le altre.
    query  = '''
        SELECT r.data, i.nome AS impianto, r.codice_pv,
               r.categoria,
               r.valore_teorico, r.valore_reale, r.differenza, r.stato, r.note, r.tipo_match
        FROM riconciliazione_risultati r
        LEFT JOIN impianti i ON r.codice_pv = i.codice_pv
        WHERE 1=1
    '''
    params = []
    if da:
        query  += " AND r.data >= ?"
        params.append(da)
    if a:
        query  += " AND r.data <= ?"
        params.append(a)
    if pv:
        query  += " AND r.codice_pv = ?"
        params.append(int(pv))
    query += " ORDER BY r.data DESC, i.nome, r.categoria"

    conn = get_connection()
    rows = conn.execute(query, params).fetchall()
    conn.close()

    # Mappa categorie interne -> etichetta leggibile in italiano
    _CAT_LABEL = {
        "carte_bancarie":          "Carte bancarie (POS)",
        "satispay":                "Satispay",
        "buoni":                   "Buoni / Voucher",
        "carte_petrolifere":       "Carte petrolifere",
        "buoni_petrolifere":       "Buoni + Petrolifere",
        "buoni_petrolifere_combined": "Buoni + Petrolifere (combinato)",
        "prove_erogazione":        "Prove di erogazione",
        "clienti_fine_mese":       "Clienti fine mese",
        "diversi":                 "Diversi",
    }

    cols = ["data", "impianto", "codice_pv", "categoria_label",
            "valore_teorico", "valore_reale", "differenza", "stato", "note", "tipo_match"]

    records = []
    for r in rows:
        cat_raw   = r["categoria"]
        cat_label = _CAT_LABEL.get(cat_raw, cat_raw)
        records.append({
            "data":           r["data"],
            "impianto":       f"{r['codice_pv']} – {r['impianto'] or 'N/D'}",
            "codice_pv":      r["codice_pv"],
            "categoria_label": cat_label,
            "valore_teorico": r["valore_teorico"],
            "valore_reale":   r["valore_reale"],
            "differenza":     r["differenza"],
            "stato":          r["stato"],
            "note":           r["note"] or "",
            "tipo_match":     r["tipo_match"] or "",
        })

    df = pd.DataFrame(records, columns=cols) if records else pd.DataFrame(columns=cols)

    df.rename(columns={
        "data":            "Data",
        "impianto":        "Impianto",
        "codice_pv":       "Cod. PV",
        "categoria_label": "Categoria",
        "valore_teorico":  "Fortech (€)",
        "valore_reale":    "Reale (€)",
        "differenza":      "Diff (€)",
        "stato":           "Stato",
        "note":            "Note",
        "tipo_match":      "Tipo Match",
    }, inplace=True)

    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Riconciliazioni")
        ws = writer.sheets["Riconciliazioni"]

        # ── Stile header ──────────────────────────────────────────────────────
        header_fill = PatternFill("solid", fgColor="1F4D49")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Colori per stato ──────────────────────────────────────────────────
        _FILL = {
            "QUADRATO":        PatternFill("solid", fgColor="C6EFCE"),  # verde
            "QUADRATO_ARROT":  PatternFill("solid", fgColor="DDEBF7"),  # azzurro
            "ANOMALIA_LIEVE":  PatternFill("solid", fgColor="FFEB9C"),  # giallo
            "ANOMALIA_GRAVE":  PatternFill("solid", fgColor="FFC7CE"),  # rosso
            "NON_TROVATO":     PatternFill("solid", fgColor="F4CCCC"),  # arancione chiaro
        }
        # Colonna "Stato" (indice 8 = colonna H nel foglio con header a riga 1)
        stato_col_idx = df.columns.get_loc("Stato") + 1
        for row_idx in range(2, ws.max_row + 1):
            stato_cell  = ws.cell(row=row_idx, column=stato_col_idx)
            stato_val   = str(stato_cell.value or "")
            fill        = _FILL.get(stato_val)
            if fill:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=c).fill = fill

        # ── Larghezza colonne automatica ──────────────────────────────────────
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

        ws.freeze_panes = "A2"

    buf.seek(0)
    fname = f"Riconciliazioni_{datetime.date.today().isoformat()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



# ═══════════════════════════════════════════════════════════════════════════════
#  CHART DATA
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/chart-data", methods=["GET"])
@jwt_required()
def get_chart_data():
    conn = get_connection()
    rows = conn.execute('''
        SELECT categoria,
               SUM(valore_teorico) AS tot_fortech,
               SUM(valore_reale)   AS tot_reale
        FROM riconciliazione_risultati
        GROUP BY categoria
        ORDER BY tot_fortech DESC
    ''').fetchall()
    conn.close()

    return jsonify([{
        "categoria":   r["categoria"],
        "tot_fortech": round(r["tot_fortech"] or 0, 2),
        "tot_reale":   round(r["tot_reale"] or 0, 2),
    } for r in rows])


# ═══════════════════════════════════════════════════════════════════════════════
#  IMPIANTI — LISTA
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/impianti", methods=["GET"])
@jwt_required()
def get_impianti():
    conn = get_connection()
    rows = conn.execute(
        "SELECT id, codice_pv, nome, comune, indirizzo, tipo_gestione FROM impianti ORDER BY nome"
    ).fetchall()

    results = []
    for r in rows:
        stats = conn.execute('''
            SELECT
                SUM(CASE WHEN stato IN ('QUADRATO','QUADRATO_ARROT') THEN 1 ELSE 0 END) AS cnt_ok,
                SUM(CASE WHEN stato = 'ANOMALIA_LIEVE'               THEN 1 ELSE 0 END) AS cnt_warn,
                SUM(CASE WHEN stato = 'ANOMALIA_GRAVE'               THEN 1 ELSE 0 END) AS cnt_grave
            FROM riconciliazione_risultati WHERE codice_pv = ?
        ''', (r["codice_pv"],)).fetchone()

        results.append({
            "id":           r["id"],
            "codice_pv":    r["codice_pv"],
            "nome":         r["nome"] or f"PV {r['codice_pv']}",
            "comune":       r["comune"],
            "indirizzo":    r["indirizzo"],
            "tipo":         r["tipo_gestione"] or "PRESIDIATO",
            "cnt_ok":       stats["cnt_ok"] or 0,
            "cnt_warn":     stats["cnt_warn"] or 0,
            "cnt_grave":    stats["cnt_grave"] or 0,
        })

    conn.close()
    return jsonify(results)


# ═══════════════════════════════════════════════════════════════════════════════
#  IMPIANTI — ANDAMENTO GIORNALIERO (per Modal)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/impianti/<int:imp_id>/andamento", methods=["GET"])
@jwt_required()
def get_andamento(imp_id):
    conn = get_connection()
    imp  = conn.execute(
        "SELECT id, codice_pv, nome, tipo_gestione FROM impianti WHERE id = ?", (imp_id,)
    ).fetchone()

    if not imp:
        conn.close()
        return jsonify(error="Impianto non trovato"), 404

    rows = conn.execute('''
        SELECT data, categoria, valore_teorico, valore_reale, differenza, stato
        FROM riconciliazione_risultati
        WHERE codice_pv = ?
        ORDER BY data DESC, categoria
    ''', (imp["codice_pv"],)).fetchall()
    conn.close()

    # Raggruppa per data
    _priority = {
        "ANOMALIA_GRAVE": 5, "ANOMALIA_LIEVE": 4, "NON_TROVATO": 3,
        "IN_ATTESA": 2, "QUADRATO_ARROT": 1, "QUADRATO": 0,
    }
    giorni_dict: dict[str, dict] = {}
    for r in rows:
        d = r["data"]
        if d not in giorni_dict:
            giorni_dict[d] = {"data": d, "categorie": {}, "totale_diff": 0.0, "stato_peggiore": "QUADRATO"}
        giorni_dict[d]["categorie"][r["categoria"]] = {
            "teorico":    float(r["valore_teorico"] or 0),
            "reale":      float(r["valore_reale"] or 0),
            "differenza": float(r["differenza"] or 0),
            "stato":      r["stato"],
        }
        giorni_dict[d]["totale_diff"] = round(
            giorni_dict[d]["totale_diff"] + float(r["differenza"] or 0), 2)
        if _priority.get(r["stato"], 0) > _priority.get(giorni_dict[d]["stato_peggiore"], 0):
            giorni_dict[d]["stato_peggiore"] = r["stato"]

    giorni = list(giorni_dict.values())

    # Statistiche aggregate
    stati_cnt = Counter(g["stato_peggiore"] for g in giorni)

    return jsonify({
        "impianto":     dict(imp),
        "totale_giorni": len(giorni),
        "stats":        dict(stati_cnt),
        "giorni":       giorni,
    })


# ═══════════════════════════════════════════════════════════════════════════════
#  SICUREZZA (Placeholder — Taleggio IoT non ancora integrato)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/sicurezza", methods=["GET"])
@jwt_required()
def get_sicurezza():
    return jsonify([])


# ═══════════════════════════════════════════════════════════════════════════════
#  AI REPORT (OpenRouter — streaming SSE)
# ═══════════════════════════════════════════════════════════════════════════════

OPENROUTER_MODELS = [
    {"id": "openai/gpt-4o-mini",                     "label": "GPT-4o Mini (veloce, economico)"},
    {"id": "openai/gpt-4o",                          "label": "GPT-4o (più potente)"},
    {"id": "anthropic/claude-3-haiku",               "label": "Claude 3 Haiku (veloce)"},
    {"id": "anthropic/claude-3.5-sonnet",            "label": "Claude 3.5 Sonnet (top Anthropic)"},
    {"id": "google/gemini-flash-1.5",                "label": "Gemini Flash 1.5 (Google)"},
    {"id": "meta-llama/llama-3.1-8b-instruct:free",  "label": "Llama 3.1 8B (gratuito)"},
]

def _build_ai_context(conn, data_from=None, data_to=None, codice_pv=None):
    """Costruisce il contesto completo per il prompt AI."""
    where_anom = ["r.stato NOT IN ('QUADRATO','QUADRATO_ARROT')"]
    where_all  = ["1=1"]
    if data_from:
        where_anom.append(f"r.data >= '{data_from}'")
        where_all.append(f"r.data >= '{data_from}'")
    if data_to:
        where_anom.append(f"r.data <= '{data_to}'")
        where_all.append(f"r.data <= '{data_to}'")
    if codice_pv:
        where_anom.append(f"r.codice_pv = {int(codice_pv)}")
        where_all.append(f"r.codice_pv = {int(codice_pv)}")

    # Anomalie ordinate per gravità
    anomalie = conn.execute(f'''
        SELECT r.data, i.nome AS impianto, r.categoria,
               r.valore_teorico, r.valore_reale, r.differenza, r.stato, r.note
        FROM riconciliazione_risultati r
        LEFT JOIN impianti i ON r.codice_pv = i.codice_pv
        WHERE {" AND ".join(where_anom)}
        ORDER BY ABS(r.differenza) DESC
        LIMIT 80
    ''').fetchall()

    if not anomalie:
        return None

    # Statistiche globali
    stats = conn.execute(f'''
        SELECT
            COUNT(*) AS totale,
            SUM(CASE WHEN stato IN ('QUADRATO','QUADRATO_ARROT') THEN 1 ELSE 0 END) AS quadrate,
            SUM(CASE WHEN stato = 'ANOMALIA_GRAVE' THEN 1 ELSE 0 END) AS gravi,
            SUM(CASE WHEN stato = 'ANOMALIA_LIEVE' THEN 1 ELSE 0 END) AS lievi,
            SUM(CASE WHEN stato = 'NON_TROVATO' THEN 1 ELSE 0 END) AS non_trovate,
            ROUND(SUM(ABS(r.differenza))::numeric, 2) AS esposizione_totale
        FROM riconciliazione_risultati r
        WHERE {" AND ".join(where_all)}
    ''').fetchone()

    # Anomalie per impianto (aggregato)
    per_impianto = conn.execute(f'''
        SELECT i.nome AS impianto,
               COUNT(*) AS n_anomalie,
               ROUND(SUM(ABS(r.differenza))::numeric, 2) AS esposizione,
               MAX(ABS(r.differenza)) AS max_diff,
               STRING_AGG(DISTINCT r.categoria, ', ') AS categorie
        FROM riconciliazione_risultati r
        LEFT JOIN impianti i ON r.codice_pv = i.codice_pv
        WHERE {" AND ".join(where_anom)}
        GROUP BY i.nome
        ORDER BY esposizione DESC
        LIMIT 15
    ''').fetchall()

    # Anomalie per categoria (aggregato)
    per_categoria = conn.execute(f'''
        SELECT r.categoria,
               COUNT(*) AS n_anomalie,
               ROUND(SUM(ABS(r.differenza))::numeric, 2) AS esposizione,
               ROUND(AVG(ABS(r.differenza))::numeric, 2) AS media_diff
        FROM riconciliazione_risultati r
        WHERE {" AND ".join(where_anom)}
        GROUP BY r.categoria
        ORDER BY esposizione DESC
    ''').fetchall()

    # Assembla contesto
    s = stats
    out = []
    out.append("=== STATISTICHE GLOBALI ===")
    out.append(f"Totale riconciliazioni: {s['totale']} | Quadrate: {s['quadrate']} | "
               f"Anomalie gravi: {s['gravi']} | Anomalie lievi: {s['lievi']} | "
               f"Non trovate: {s['non_trovate']} | Esposizione totale: €{s['esposizione_totale']}")

    out.append("\n=== ANOMALIE PER IMPIANTO (aggregate) ===")
    out.append("Impianto | N.Anomalie | Esposizione€ | MaxDiff€ | Categorie")
    for r in per_impianto:
        out.append(f"{r['impianto'] or 'N/D'} | {r['n_anomalie']} | {r['esposizione']} | {r['max_diff']:.2f} | {r['categorie']}")

    out.append("\n=== ANOMALIE PER CATEGORIA (aggregate) ===")
    out.append("Categoria | N.Anomalie | Esposizione€ | MediaDiff€")
    for r in per_categoria:
        out.append(f"{r['categoria']} | {r['n_anomalie']} | {r['esposizione']} | {r['media_diff']}")

    out.append("\n=== DETTAGLIO ANOMALIE (ordinate per gravità) ===")
    out.append("Data | Impianto | Categoria | Fortech€ | Reale€ | Diff€ | Stato | Note")
    for r in anomalie:
        note = (r['note'] or '').replace('\n', ' ')
        out.append(
            f"{r['data']} | {r['impianto'] or 'N/D'} | {r['categoria']} | "
            f"{float(r['valore_teorico']):.2f} | {float(r['valore_reale']):.2f} | "
            f"{float(r['differenza']):.2f} | {r['stato']} | {note}"
        )

    return "\n".join(out)


def _build_prompt(context):
    return f"""Sei un controller finanziario specializzato nella riconciliazione contabile di stazioni di servizio carburante (benzinai) in Italia. Hai accesso ai dati di riconciliazione tra i valori teorici del sistema gestionale (Fortech) e i valori reali incassati tramite POS, contanti, Satispay, buoni e carte petrolifere.

GLOSSARIO DEI TIPI DI ANOMALIA:
- ANOMALIA_GRAVE: differenza > €50 tra teorico e reale — richiede intervento immediato
- ANOMALIA_LIEVE: differenza tra €1 e €50 — da monitorare
- NON_TROVATO: il sistema Fortech registra un incasso ma non esiste la transazione reale corrispondente
- QUADRATO_ARROT: quadrato entro tolleranza di arrotondamento (ok)

CATEGORIE DI INCASSO:
- pos/carte_bancarie: transazioni POS (Visa, Mastercard, Bancomat)
- satispay: pagamenti digitali Satispay
- buoni: voucher e buoni carburante
- petrolifere/carte_petrolifere: DKV, UTA, CartaMaxima
- buoni_petrolifere_combined: impianti senza separazione tra buoni e carte petrolifere
- contanti: versamenti contanti (attualmente non riconciliati automaticamente)

DATI DI RICONCILIAZIONE:
{context}

---

Genera un report professionale completo in italiano con le seguenti sezioni obbligatorie:

## 1. Sintesi Esecutiva
Panoramica in 3-5 righe: quante anomalie, esposizione totale, impianti più critici, trend generale.

## 2. Anomalie Critiche per Impianto
Per ogni impianto con anomalie GRAVI o NON_TROVATO, analizza:
- Quale categoria è problematica e perché
- Entità della discrepanza in euro
- Se il problema è isolato (singola data) o sistematico (più date)
- Giudizio sul rischio: ALTO / MEDIO / BASSO

## 3. Analisi per Categoria di Incasso
Per ogni categoria con anomalie significative:
- Numero anomalie e esposizione totale
- Possibili cause tecniche o operative (es: terminale POS non sincronizzato, file non caricato, errore di mappatura)
- Se il problema riguarda uno o più impianti specifici

## 4. Azioni Consigliate
Lista prioritizzata di azioni concrete (usa ✅ / ⚠️ / 🔴 per priorità):
- Chi deve fare cosa
- Entro quando
- Come verificare la risoluzione

## 5. Impianti da Monitorare
Lista degli impianti che richiedono attenzione nelle prossime riconciliazioni, con motivazione.

Usa un linguaggio professionale ma diretto. Includi i valori in euro dove rilevante. Sii specifico sui nomi degli impianti e delle categorie."""


@app.route("/api/ai-report/models", methods=["GET"])
@jwt_required()
def ai_report_models():
    conn = get_connection()
    cfg  = get_config(conn)
    conn.close()
    current = cfg.get("openrouter_model", "openai/gpt-4o-mini")
    return jsonify(models=OPENROUTER_MODELS, current=current)


@app.route("/api/ai-report/model", methods=["POST"])
@jwt_required()
def ai_report_set_model():
    body  = request.get_json(force=True) or {}
    model = body.get("model", "").strip()
    valid_ids = {m["id"] for m in OPENROUTER_MODELS}
    if model not in valid_ids:
        return jsonify(error="Modello non valido"), 400
    conn = get_connection()
    conn.execute(
        "INSERT INTO config (chiave, valore) VALUES ('openrouter_model', ?) "
        "ON CONFLICT(chiave) DO UPDATE SET valore = excluded.valore", (model,)
    )
    conn.commit()
    conn.close()
    return jsonify(message="Modello salvato")


@app.route("/api/ai-report/stream", methods=["POST"])
@jwt_required()
def ai_report_stream():
    import requests as req_lib
    from flask import stream_with_context, Response

    body = request.get_json(force=True, silent=True) or {}

    conn = get_connection()
    cfg  = get_config(conn)
    api_key = cfg.get("openrouter_api_key", "").strip()
    model   = body.get("model") or cfg.get("openrouter_model", "openai/gpt-4o-mini")

    if not api_key:
        conn.close()
        return jsonify(error="Chiave API OpenRouter non configurata. Vai in Impostazioni."), 400

    context = _build_ai_context(
        conn,
        data_from  = body.get("data_from"),
        data_to    = body.get("data_to"),
        codice_pv  = body.get("codice_pv"),
    )
    conn.close()

    if not context:
        def _empty():
            yield "data: {\"content\": \"✅ Nessuna anomalia rilevata. Tutti i valori sono quadrati.\"}\n\n"
            yield "data: [DONE]\n\n"
        return Response(stream_with_context(_empty()), mimetype="text/event-stream",
                        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

    prompt = _build_prompt(context)

    def generate():
        try:
            resp = req_lib.post(
                "https://openrouter.ai/api/v1/chat/completions",
                json={
                    "model":    model,
                    "messages": [{"role": "user", "content": prompt}],
                    "max_tokens": 4000,
                    "stream":   True,
                },
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type":  "application/json",
                    "HTTP-Referer":  "https://calor-systems.local",
                },
                stream=True,
                timeout=60,
            )
            for line in resp.iter_lines():
                if not line:
                    continue
                if isinstance(line, bytes):
                    line = line.decode("utf-8")
                if not line.startswith("data: "):
                    continue
                data_str = line[6:]
                if data_str.strip() == "[DONE]":
                    yield "data: [DONE]\n\n"
                    return
                try:
                    chunk   = json.loads(data_str)
                    content = chunk["choices"][0]["delta"].get("content", "")
                    if content:
                        yield f"data: {json.dumps({'content': content})}\n\n"
                except Exception:
                    pass
        except Exception as e:
            yield f"data: {{\"error\": \"{str(e)}\"}}\n\n"
            yield "data: [DONE]\n\n"

    return Response(stream_with_context(generate()), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/api/ai-report", methods=["POST"])
@jwt_required()
def ai_report():
    # Endpoint legacy — redirige allo stream per retrocompatibilità
    return ai_report_stream()



# ═══════════════════════════════════════════════════════════════════════════════
#  IMPOSTAZIONI — CONFIG
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/settings/config", methods=["GET", "POST"])
@jwt_required()
def settings_config():
    conn = get_connection()
    if request.method == "GET":
        cfg = get_config(conn)
        conn.close()
        # Ritorna solo le chiavi non sensibili
        safe_keys = [
            "tolleranza_carte_fisiologica", "tolleranza_satispay", "scarto_giorni_buoni",
        ]
        return jsonify({k: float(cfg[k]) if k in cfg else 0 for k in safe_keys})

    data = request.get_json(force=True)
    for k, v in data.items():
        conn.execute(
            "INSERT INTO config (chiave, valore) VALUES (?, ?) "
            "ON CONFLICT(chiave) DO UPDATE SET valore = excluded.valore",
            (k, str(v))
        )
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  IMPOSTAZIONI — PASSWORD
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/settings/password", methods=["POST"])
@jwt_required()
def settings_password():
    identity = get_jwt_identity()
    data     = request.get_json(force=True)
    old_pw   = data.get("old_password", "")
    new_pw   = data.get("new_password", "")

    if len(new_pw) < 8:
        return jsonify(msg="La nuova password deve avere almeno 8 caratteri"), 400

    conn = get_connection()
    row  = conn.execute(
        "SELECT password_hash FROM users WHERE username = ?", (identity,)
    ).fetchone()

    if not row or row["password_hash"] != hashlib.sha256(old_pw.encode()).hexdigest():
        conn.close()
        return jsonify(msg="Vecchia password non corretta"), 400

    new_hash = hashlib.sha256(new_pw.encode()).hexdigest()
    conn.execute(
        "UPDATE users SET password_hash = ? WHERE username = ?", (new_hash, identity)
    )
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  IMPOSTAZIONI — API KEY
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/settings/apikey", methods=["GET", "POST"])
@jwt_required()
def settings_apikey():
    conn = get_connection()
    if request.method == "GET":
        cfg = get_config(conn)
        conn.close()
        key = cfg.get("openrouter_api_key", "")
        if key:
            masked = key[:8] + "…" + key[-4:]
            return jsonify(has_key=True, masked=masked)
        return jsonify(has_key=False, masked="")

    data    = request.get_json(force=True)
    api_key = data.get("api_key", "").strip()
    if not api_key:
        conn.close()
        return jsonify(error="Chiave vuota"), 400

    conn.execute(
        "INSERT INTO config (chiave, valore) VALUES ('openrouter_api_key', ?) "
        "ON CONFLICT(chiave) DO UPDATE SET valore = excluded.valore", (api_key,)
    )
    conn.commit()
    conn.close()
    return jsonify(message="Chiave API salvata con successo")


@app.route("/api/settings/apikey/test", methods=["POST"])
@jwt_required()
def settings_apikey_test():
    import requests as req_lib

    data    = request.get_json(force=True)
    api_key = data.get("api_key", "").strip()

    if not api_key:
        conn = get_connection()
        cfg  = get_config(conn)
        conn.close()
        api_key = cfg.get("openrouter_api_key", "").strip()

    if not api_key:
        return jsonify(error="Nessuna chiave API configurata"), 400

    try:
        resp = req_lib.post(
            "https://openrouter.ai/api/v1/chat/completions",
            json={"model": "openai/gpt-4o-mini", "messages": [{"role": "user", "content": "ping"}], "max_tokens": 5},
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            timeout=10,
        )
        if resp.status_code == 401:
            return jsonify(error="❌ Chiave non valida (401 Unauthorized)"), 400
        if not resp.ok:
            return jsonify(error=f"❌ HTTP {resp.status_code}"), 502
        return jsonify(message="✅ Connessione OpenRouter OK")
    except Exception as e:
        return jsonify(error=f"❌ Connessione fallita: {e}"), 502


# ═══════════════════════════════════════════════════════════════════════════════
#  AVVIO
# ═══════════════════════════════════════════════════════════════════════════════
#  VERIFICA COMPENSAZIONI
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/verifica/anteprima", methods=["POST"])
@jwt_required()
def verifica_anteprima():
    """Restituisce le coppie compensabili senza modificare il DB."""
    from verifica import find_compensazioni
    body       = request.get_json(force=True, silent=True) or {}
    tolleranza = float(body.get("tolleranza", 1.0))
    codice_pv  = body.get("codice_pv") or None
    categoria  = body.get("categoria") or None

    conn    = get_connection()
    matches = find_compensazioni(conn, tolleranza=tolleranza, codice_pv=codice_pv, categoria=categoria)
    conn.close()

    # Statistiche riassuntive
    esposizione_totale = sum(abs(m['diff_pos']) for m in matches)
    return jsonify(
        matches=matches,
        totale_coppie=len(matches),
        righe_compensabili=len(matches) * 2,
        esposizione_compensata=round(esposizione_totale, 2),
    )


@app.route("/api/verifica/applica", methods=["POST"])
@jwt_required()
def verifica_applica():
    """Applica tutte le compensazioni (o un sottoinsieme tramite id_pairs)."""
    from verifica import applica_compensazioni
    body       = request.get_json(force=True, silent=True) or {}
    tolleranza = float(body.get("tolleranza", 1.0))
    id_pairs   = body.get("id_pairs")  # lista di [id_pos, id_neg] o None

    if id_pairs is not None:
        id_pairs = [tuple(p) for p in id_pairs]

    conn = get_connection()
    updated, matches = applica_compensazioni(conn, id_pairs=id_pairs, tolleranza=tolleranza)
    conn.close()

    return jsonify(
        message=f"{updated} righe aggiornate a QUADRATO_COMPENSATO",
        coppie_compensate=len(matches),
        righe_aggiornate=updated,
    )


@app.route("/api/verifica/reset", methods=["POST"])
@jwt_required()
def verifica_reset():
    """Annulla tutte le compensazioni (ripristina ANOMALIA_GRAVE)."""
    from verifica import reset_compensazioni
    body      = request.get_json(force=True, silent=True) or {}
    codice_pv = body.get("codice_pv") or None

    conn = get_connection()
    reset_compensazioni(conn, codice_pv=codice_pv)
    conn.close()
    return jsonify(message="Compensazioni annullate.")


@app.route("/api/verifica/stats", methods=["GET"])
@jwt_required()
def verifica_stats():
    """Statistiche sullo stato delle compensazioni."""
    conn = get_connection()
    rows = conn.execute("""
        SELECT stato, COUNT(*) as n, ROUND(SUM(ABS(differenza))::numeric, 2) as esposizione
        FROM riconciliazione_risultati
        GROUP BY stato
    """).fetchall()
    conn.close()
    return jsonify({r['stato']: {'count': r['n'], 'esposizione': float(r['esposizione'] or 0)} for r in rows})


# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    from database import init_db
    init_db()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)

